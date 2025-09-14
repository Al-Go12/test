from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from .models import *
from accounts.serializers import * 
from django.db.models import Count
from django.utils.timezone import now 
from django.db.models.functions import ExtractMonth 
from django.db.models import Q
from django.utils.dateparse import parse_date 
from django.http import HttpResponse 
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment  
import requests
from openpyxl.utils import get_column_letter
import json, uuid
# accounts/views.py
class ProfileAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        profile, created = Profile.objects.get_or_create(user=request.user)
        serializer = ProfileSerializer(profile)
        return Response(serializer.data, status=status.HTTP_200_OK)


    def post(self, request):
        """Create or update profile; create registration only if needed"""
        profile, created = Profile.objects.get_or_create(user=request.user)
        serializer = ProfileSerializer(profile, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save(user=request.user)

            # Only create registration if not exists and required
            registration = Registration.objects.filter(user=request.user).first()
            if not registration:
                registration = Registration.objects.create(user=request.user, profile=profile)

            return Response({
                "profile": serializer.data,
                "registration": {
                    "id": registration.id,
                    "status": registration.status,
                    "submitted_at": registration.submitted_at,
                }
            }, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class RegistrationStatusAPIView(APIView):
    """
    API to check registration status for authenticated user
    GET /api/registration/status/
    """ 
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request): 
        try:
            user = request.user  # Even simpler - no need for pk lookup
            
            # Check if registration exists for this user
            try:
                registration = Registration.objects.select_related('user', 'profile').get(user=user)
                
                serializer = RegistrationStatusSerializer(registration)
                
                return Response({
                    'exists': True,
                    'message': 'Registration found',
                    'application': serializer.data
                }, status=status.HTTP_200_OK)
                
            except Registration.DoesNotExist:
                return Response({
                    'exists': False,
                    'message': 'User exists but no registration found',
                    'application': None,
                    'user_info': {
                        'phone_number':  str(user.phone_number),
                        'name': getattr(user, 'name', ''),
                    }
                }, status=status.HTTP_200_OK)
                
        except Exception as e:
            return Response({
                'exists': False,
                'message': 'An error occurred while checking registration status',
                'error': str(e),
                'application': None
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)







class UserApplicationAPIView(APIView):
    """
    Get the latest registration/application of the authenticated user
    GET /api/registration/application/
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        try:
            registration = Registration.objects.filter(user=request.user).latest('submitted_at')
            serializer = RegistrationSerializer(registration)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Registration.DoesNotExist:
            return Response(
                {"status":False,"error": "No application found"},
                status=status.HTTP_404_NOT_FOUND
            )


































class AdminProfileListView(APIView):
    permission_classes = [permissions.IsAuthenticated]  # ✅ only logged-in users (admins)

    def get(self, request):
        profiles = Profile.objects.all().order_by("-id")

        # --- 🔎 Search Filter ---
        search = request.query_params.get("search")
        if search:
            profiles = profiles.filter(
                Q(full_name__icontains=search) |
                Q(contact_number__icontains=search) |
                Q(alternative_contact_number__icontains=search)
            )

        # --- 📅 Date Filter ---
        from_date = request.query_params.get("from_date")
        to_date = request.query_params.get("to_date")

        if from_date:
            profiles = profiles.filter(created_at__date__gte=parse_date(from_date))
        if to_date:
            profiles = profiles.filter(created_at__date__lte=parse_date(to_date))

        # --- 📄 Pagination ---
        page = int(request.query_params.get("page", 1))
        page_size = int(request.query_params.get("page_size", 10))
        start = (page - 1) * page_size
        end = start + page_size

        total_items = profiles.count()
        total_pages = (total_items + page_size - 1) // page_size

        serializer = ProfileSerializer(profiles[start:end], many=True)

        return Response({
            "results": serializer.data,
            "pagination": {
                "current_page": page,
                "page_size": page_size,
                "total_items": total_items,
                "total_pages": total_pages,
            }
        }, status=status.HTTP_200_OK)


class DashboardAPIView(APIView): 
    permission_classes = [permissions.IsAuthenticated] 

    def get(self, request):
        # 1. Total applicants
        total_applicants = Registration.objects.count()

        # 2. Last 5 registrations
        last_five = Registration.objects.select_related("profile").order_by("-submitted_at")[:5]
        last_five_data = RegistrationSerializer(last_five, many=True).data

        # 3. Registrations by status (for donut chart)
        status_counts = (
            Registration.objects.values("status")
            .annotate(count=Count("id"))
            .order_by()
        )
        donut_data = {item["status"]: item["count"] for item in status_counts}

        # 4. Registrations per month (for bar chart)
        current_year = now().year
        monthly_counts = (
    Registration.objects.filter(submitted_at__year=current_year)
    .annotate(month=ExtractMonth("submitted_at"))
    .values("month")
    .annotate(count=Count("id"))
    .order_by("month")
)
        bar_data = []
        for m in range(1, 13):
            count = next((x["count"] for x in monthly_counts if int(x["month"]) == m), 0)
            bar_data.append({"month": m, "count": count})

        # Final response
        data = {
            "total_applicants": total_applicants,
            "last_five_registrations": last_five_data,
            "donut_chart": donut_data,
            "bar_chart": bar_data,
        }

        return Response(data, status=status.HTTP_200_OK)         







class AdminProfileExportExcelView(APIView):
    permission_classes = [permissions.IsAuthenticated]  # Only admins/logged-in users

    def get(self, request):
        from_date = request.query_params.get("from_date")
        to_date = request.query_params.get("to_date")

        profiles = Profile.objects.all().order_by("-id")
        if from_date:
            profiles = profiles.filter(created_at__date__gte=parse_date(from_date))
        if to_date:
            profiles = profiles.filter(created_at__date__lte=parse_date(to_date))

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Profiles"

        # Header row
        headers = [
             "Full Name", "Loan Type","Loan Amount","Aadhaar Number", "PAN Number", "Bank Name",
            "Account Number", "IFSC Code", "Full Address", "Contact Number",
            "Alternative Contact Number", "City", "PIN Code", "State"
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_fill_color = "4F81BD"  # nice blue header

        # Add headers with styling
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = openpyxl.styles.PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")

        # Set column widths
        column_widths = [6,10,10,20, 18, 14, 18, 20, 14, 40, 15, 18, 15, 10, 15,]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Data rows
        for row_idx, p in enumerate(profiles, 2):
           
            ws.cell(row=row_idx, column=1, value=p.full_name) 
            ws.cell(row=row_idx, column=2, value=p.loan_type)
            ws.cell(row=row_idx, column=3, value=p.loan_amount)
            ws.cell(row=row_idx, column=4, value=p.aadhaar_number)
            ws.cell(row=row_idx, column=5, value=p.pan_number)
            ws.cell(row=row_idx, column=6, value=p.bank_name)
            ws.cell(row=row_idx, column=7, value=p.account_number)
            ws.cell(row=row_idx, column=8, value=p.ifsc_code)
            ws.cell(row=row_idx, column=9, value=p.full_address)
            ws.cell(row=row_idx, column=10, value=p.contact_number or "")
            ws.cell(row=row_idx, column=11, value=p.alternative_contact_number or "")
            ws.cell(row=row_idx, column=12, value=p.city)
            ws.cell(row=row_idx, column=13, value=p.pin_code)
            ws.cell(row=row_idx, column=14, value=p.state)
          
            # Optional: Align all text left except IDs and dates
            for col in range(1, 16):
                cell = ws.cell(row=row_idx, column=col)
                if col in [1, 14, 15]:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

            # Optional: Adjust row height
            ws.row_dimensions[row_idx].height = 18

        # Return as Excel file
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        filename = f"Profiles_{from_date or 'start'}_to_{to_date or 'end'}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response          

 
def cf_headers():
  return {
    "x-client-id": settings.CASHFREE_APP_ID,
    "x-client-secret": settings.CASHFREE_SECRET_KEY,
    "x-api-version": settings.CASHFREE_API_VERSION,
    "Content-Type": "application/json",
  }

class CashfreeCreateOrderView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        profile = getattr(user, "profile", None) 
        
        if not profile:
            return Response({"error": "Profile not found"}, status=status.HTTP_400_BAD_REQUEST)

        amount = 100  # INR as requested
        order_id = f"order_{uuid.uuid4().hex[:16]}" 
        registration=Registration.objects.get(user=user)

        payload = {
            "order_id": order_id,
            "order_amount": float(amount),
            "order_currency": "INR",
            "customer_details": {
                "customer_id": str(registration.id),
                "customer_name": profile.full_name or user.get_username(),
                "customer_email":  "example@gmail.com",
                "customer_phone": profile.contact_number or "9999999999",
            },
            "order_meta": {
                # optional but recommended: your return and webhook URLs
                # "return_url": f"https://yourapp.com/return?order_id={order_id}",
                # "notify_url": "https://yourapp.com/api/payments/cashfree-webhook",
            },
        }

        r = requests.post(f"{settings.CASHFREE_PG_BASE}/orders", headers=cf_headers(), data=json.dumps(payload), timeout=15)
        data = r.json()
        if r.status_code >= 300:
            return Response(data, status=r.status_code)

        # Prefer payment_session_id from create order; fallback to sessions API if needed
        payment_session_id = data.get("payment_session_id")
        if not payment_session_id:
            s = requests.post(f"{settings.CASHFREE_PG_BASE}/orders/sessions", headers=cf_headers(), data=json.dumps({"order_id": order_id}), timeout=15)
            sd = s.json()
            if s.status_code >= 300:
                return Response(sd, status=s.status_code)
            payment_session_id = sd.get("payment_session_id")

        return Response({"order_id": order_id, "payment_session_id": payment_session_id})      
    

class CashfreePaymentStatusView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    

    def get(self, request, order_id):
        if not order_id:
            return Response({"error": "Missing order_id parameter"}, status=status.HTTP_400_BAD_REQUEST)

        url = f"{settings.CASHFREE_PG_BASE}/orders/{order_id}/payments"
        headers = {
            "x-client-id": settings.CASHFREE_APP_ID,
            "x-client-secret": settings.CASHFREE_SECRET_KEY,
            "x-api-version": settings.CASHFREE_API_VERSION,
            "Content-Type": "application/json",
        }

        try:
            r = requests.get(url, headers=headers, timeout=15)
            payment_data = r.json()
        except Exception as e:
            return Response({"error": f"Failed to fetch payment status: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        if r.status_code >= 300:
            return Response(payment_data, status=r.status_code)

        if not payment_data:
            return Response({"status": "PENDING", "payment_data": payment_data})

        if isinstance(payment_data, list): 
            print(payment_data)
            first_payment = payment_data[0] if len(payment_data) > 0 else {}
            status_value = first_payment.get("payment_status") or first_payment.get("status") or "PENDING"
        else:
            status_value = payment_data.get("order_status") or payment_data.get("data", {}).get("order_status") or "PENDING"

        if status_value in ["PAID", "SUCCESS", "COMPLETED"]:     
            registration=Registration.objects.get(user=request.user)  
            registration.status='paid' 
            registration.save()
            return Response({"status": status_value, "payment_data": payment_data})

        return Response({"status": status_value, "payment_data": payment_data})


class CashfreePaymentWebhookView(APIView):
    permission_classes = [permissions.AllowAny]  # Called by Cashfree

    def post(self, request):
        data = request.data
        order_id = data.get("orderId")
        txn_status = data.get("txStatus")  # SUCCESS / FAILED
        txn_amount = data.get("orderAmount")
        payment_id = data.get("referenceId")
        signature = data.get("signature")

        try:
            registration = Registration.objects.get(id__exact=order_id.split("CFO")[1])  
        except Registration.DoesNotExist:
            return Response({"error": "Registration not found"}, status=400)

        if txn_status == "SUCCESS":
            Payment.objects.create(
                registration=registration,
                gateway="cashfree",
                order_id=order_id,
                payment_id=payment_id,
                amount=int(float(txn_amount)*100),  # Convert to paise
                status="captured",
                currency="INR",
                signature=signature,
                raw_payload=data,
                webhook_received_at=timezone.now()
            ) 
            registration.status='paid' 
            registration.save()
            return Response({"status": "success"})
        else:
            Payment.objects.create(
                registration=registration,
                gateway="cashfree",
                order_id=order_id,
                payment_id=payment_id,
                amount=int(float(txn_amount)*100),
                status="failed",
                currency="INR",
                raw_payload=data,
                webhook_received_at=timezone.now()
            )
            return Response({"status": "failed"})             
        



